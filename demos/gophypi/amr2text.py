"""
Created on 4 sept. 2017

@author: Guy Lapalme <lapalme@iro.umontreal.ca> / Université de Montréal
"""
import re,datetime,os,sys

from SemanticRep import SemanticRep
import SemR2SyntR
from dereification import dereify

# from jsRealBclass  import S,Q, jsRealB
from pyrealb import *
from levenshtein import compareLevenshtein

from openpyxl import Workbook
from openpyxl.styles import Alignment

linkPat=r'<a[^>]+>(.*?)</a>'
unknownPat=r'\[\[(.*?)\]\]'
def clean(ins):
    (ins,n)=re.subn(linkPat,r'\1',ins)
    while n>0: # take into account nested links
        (ins,n)=re.subn(linkPat,r'\1',ins)
    return re.sub(unknownPat,r'\1',ins)

# HACK: some global variables...
amrs=[]
sntIn=None
sntOut=None

def amr2text(amrString,trace=False):
    if trace:print("*** AMR\n%s"%amrString)
    amrs.append(amrString)
    semR=SemanticRep.fromString(amrString)
    if trace:
        print("*** Semantic Representation - with inverse roles\n%s"%semR.prettyStr())
    dereified=dereify(semR)
    if dereified is not None:
        semR=dereified
        if trace:
            print("*** Semantic Representation - [dereified]\n%s"%semR.prettyStr())
    semR.elimInv() # transform inverse roles
    if trace:
        print("*** Semantic Representation\n%s"%semR.prettyStr())
#         print("*** Baseline Generation\n%s"%semR.baselineGen())    
    syntR=SemR2SyntR.makeSyntR(semR)
    if syntR is str:
        return syntR
    else:
        if not isinstance(syntR, (S,Q)):
            syntR=S(syntR) # ensure that the complete output is a full sentence
        if trace:
            print("*** Syntactic Representation")
            print(syntR.toSource(0))
        return syntR.realize()

def processFile(fileName,regex=None,trace=False,saveInOut=False):
    if trace:
        print("*** processFile:%s"%fileName)
    if regex is not None:
        print("*** regex:%s"%regex)
    sent=""
    amrStr=""
    amrId=0
    no=0
    for line in open(fileName,encoding='utf-8'):
        line=line.rstrip()
        if len(line)==0: ## traiter l'AMR
            if len(amrStr)==0:continue
            if regex is None or re.search(regex,amrStr,re.DOTALL):
                no+=1
                print("[%d] id:%s"%(no,amrId))
                print("sent:%s"%sent)
                if saveInOut:sntIn.write(sent+"\n")
                if not trace:print(amrStr)
                out=amr2text(amrStr,trace)
                print(out)
                if saveInOut:
                    sntOut.write(clean(out)+"\n")
                print('--')
            amrStr=""
            sent=""
        elif line[0]=="#":
            m=re.match(r'# +::id (\S+)',line)
            if m: amrId=m.group(1)
            else:
                m=re.match(r'# +::snt (.*)',line)
                if m and sent=="": # ne garder que la première phrase
                    sent=m.group(1).strip()
                else: 
                    continue
        else: ## accumuler l'AMR
            amrStr+=line+"\n"
    print("--- end of file:%s"%fileName)

def c(ws,r,c,v,style=None,nbFmt=None):
    cell=ws.cell(row=r,column=c,value=v)
    cell.alignment = Alignment(wrapText=True,vertical="top")
    if style is not None:
        cell.style=style
    if nbFmt is not None:
        cell.number_format=nbFmt
  
def createExcelWB(orig,generated,gophiPyBleu,spring=None,springBleu=None):
    wb = Workbook()
    ws = wb.active
    c(ws,1,1,"AMR")
    c(ws,1,2,"base gen")
    c(ws,1,3,"original")
    c(ws,1,4,"gophiPy")
    c(ws,1,5,"score")
    c(ws,1,6,"comment")
    if spring is not None:
        c(ws,1,7,"SPRING")
        c(ws,1,8,"score")
        c(ws,1,9,"comment")
    ws.column_dimensions["A"].width=80
    ws.column_dimensions["B"].width=30
    ws.column_dimensions["C"].width=30
    ws.column_dimensions["D"].width=30
    ws.column_dimensions["G"].width=30
    r=2
    for amr in amrs:
        c(ws,r,1,amr.strip())
        semR=SemanticRep.fromString(amr)
        c(ws,r,2,semR.baselineGen())
        c(ws,r,3,orig.readline().strip())
        c(ws,r,4,generated.readline().strip())
        c(ws,r,5,"") # ensure alignement...
        c(ws,r,6,"")
        if spring is not None:
            c(ws,r,7,spring.readline().strip())
            c(ws,r,8,"")
            c(ws,r,9,"")
        r+=1
    last=r-1
    r+=1
    c(ws,r,1,"BLEU score")
    c(ws,r,4,gophiPyBleu,nbFmt="0.00")
    c(ws,r,5,"=AVERAGE(E2:E%s)"%last,nbFmt="0.00")
    if spring is not None:
        c(ws,r,7,springBleu,nbFmt="0.00")
        c(ws,r,8,"=AVERAGE(H2:H%s)"%last,nbFmt="0.00")
    
    notes=["Generator error","Charabia/Bad info","Barely understandable","OK, but bad syntax","Perfect"]
    for i in range(4,-1,-1):
        r+=1
        c(ws,r,3,notes[i])
        c(ws,r,4,i)
        c(ws,r,5,"=COUNTIF(E$2:E$%d,D%d)"%(last,r))
        c(ws,r,6,"=E%d/E%d"%(r,last+8),"Percent","0%")
        if spring is not None:
            c(ws,r,8,"=COUNTIF(H$2:H$%d,D%d)"%(last,r))
            c(ws,r,9,"=H%d/H%d"%(r,last+8),"Percent","0%")
    r+=1
    c(ws,r,5,"=SUM(E%d:E%d)"%(r-5,r-1))
    c(ws,r,6,"=E%d/ROWS(E2:E%d)"%(r,last),"Percent","0%")
    if spring is not None:
        c(ws,r,8,"=SUM(H%d:H%d)"%(r-5,r-1))
        c(ws,r,9,"=H%d/ROWS(H2:H%d)"%(r,last),"Percent","0%")
    r+=1
    dateTime=datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
#     print("%d,%d=%s"%(r,1,dateTime))
    c(ws,r,1,dateTime)
    ws.column_dimensions["B"].hidden=True       # cacher le baseline
    ws.freeze_panes=ws["A2"]                    # figer la première ligne
    ws.auto_filter.ref=("A1:J%d"%last)          # ajouter les filtres sur la première ligne
    return wb    

def showAMRsFile(fileName,regex=r"",trace=False,createExcel=False):
    global sntIn,sntOut
    saveInOut = len(regex)==0
    ## file creation to compute BLEU
    if saveInOut:
        sntIn=open(resultDir+"sntIn.txt","w",encoding='utf-8')
        sntOut=open(resultDir+"sntOut.txt","w",encoding='utf-8')
    processFile(fileName,regex,trace,saveInOut)
    if saveInOut:
        from calculatebleu import BLEU,fetch_data
        sntIn.close()
        sntOut.close()
        candidate, references = fetch_data(resultDir+"sntOut.txt",resultDir+"sntIn.txt")
        bleuGophiPy=BLEU(candidate,references)
        orig=open(resultDir+"sntIn.txt",encoding="utf-8")
        generated=open(resultDir+"sntOut.txt",encoding="utf-8")
        for (o,g) in zip(orig.readlines(),generated.readlines()):
            print("%d: %s"%compareLevenshtein(o.strip(),g.strip()))
        print("%d examples, BLEU:%5.3f"%(len(candidate),bleuGophiPy))
        if createExcel:
            orig=open(resultDir+"sntIn.txt",encoding="utf-8")
            generated=open(resultDir+"sntOut.txt",encoding="utf-8")
            try :
                springFile=open(fileName.replace(".txt","-SPRING.out"),encoding="utf-8")
                bleuSpring=34.62
                wb = createExcelWB(orig, generated, bleuGophiPy, springFile, bleuSpring)
            except IOError:
                wb=createExcelWB(orig,generated,bleuGophiPy)
            xlsxFile= fileName.replace(".txt",".xlsx")      
            wb.save(xlsxFile)
            print(xlsxFile+" written")        
    
def init_pyrealb():
    loadEn()
    addToLexicon("tsunami",{"N":{"tab":"n1"}})

dataDir="/Users/lapalme/Dropbox/AMR/"
resultDir=os.path.join(sys.path[0],"Results/")
## Files with examples
#     fileName=dataDir+"amr-release-2.0-amrs-training-guidelines.txt"
# fileName=dataDir+"abstract_meaning_representation_amr_2.0/data/amrs/unsplit/amr-release-2.0-amrs-bolt.txt"
#     fileName=dataDir+"inverse-relations.txt"
#     fileName=dataDir+"amr-examples/amr-reifications-examples.txt"
#     fileName=dataDir+"amr-bank-PetitPrince-dev.txt"
fileName=dataDir+"amr-examples/amr-examples.txt"
#     fileName="/Users/lapalme/Dropbox/AMR2en/data/morph-examples.txt"   
# fileName=dataDir+"amr-examples/amr-dict-examples.txt"
# fileName=dataDir+"amr-examples/amr-guidelines-1_2_5.txt"
# fileName=dataDir+"amr_annotation_3.0/data/amrs/unsplit/amr-release-3.0-amrs-dfa.txt"

# fileName=dataDir+"amr_annotation_3.0/data/amrs/split/test/amr-release-3.0-amrs-test-sample-25-5.txt"
if __name__ == '__main__': # examples of call
    init_pyrealb()
    showAMRsFile(fileName,r"",False,False)

# print(amr2text('''
# (m / multi-sentence
#       :snt1 (a / and
#             :op1 (h / have-03
#                   :ARG0 (p / person
#                         :ARG1-of (s3 / star-01))
#                   :ARG1 (c3 / contract-02
#                         :quant (l / lot)))
#             :op2 (r / roll-01
#                   :ARG1 (s2 / source-01
#                         :ARG0 p
#                         :ARG1 (m2 / money))
#                   :condition (t2 / tackle-01
#                         :ARG0 p
#                         :manner (w / well-09))))
#       :snt2 (c2 / come-02
#             :ARG0 (o2 / or
#                   :op1 (f2 / force-01
#                         :ARG2 (o / out))
#                   :op2 (f3 / frame-02
#                         :mod (e / even)))
#             :ARG1 (h2 / heel
#                   :part-of (s / she))
#             :condition (t3 / tack-00 :polarity -
#                   :ARG0 s
#                   :ARG1-of (w2 / well-05))))
# ''', True))
